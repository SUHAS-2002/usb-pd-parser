from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from src.core.base_report_generator import BaseReportGenerator
from src.utils.jsonl_utils import JSONLHandler
from src.config import CONFIG


class ExcelReportGenerator(BaseReportGenerator):
    """
    Excel coverage report generator.

    Encapsulation:
    - ALL internal state uses name-mangled attributes (__attr)
    - Public API inherited from BaseReportGenerator
    """

    # --------------------------------------------------------------
    # TRUE PRIVATE class constants
    # --------------------------------------------------------------
    __STATUS_MATCHED = "MATCHED"
    __STATUS_MISSING = "MISSING_IN_CONTENT"
    __STATUS_EXTRA = "EXTRA_IN_CONTENT"

    # --------------------------------------------------------------
    # Constructor
    # --------------------------------------------------------------
    def __init__(
        self,
        toc_path: str,
        chunks_path: str,
        output_xlsx: str,
    ) -> None:
        super().__init__(output_xlsx)

        self.__toc_path: Path = Path(toc_path)
        self.__chunks_path: Path = Path(chunks_path)

        self.__colors: Dict[str, str] = {
            self.__STATUS_MATCHED: CONFIG.excel_colors.GREEN,
            self.__STATUS_MISSING: CONFIG.excel_colors.RED,
            self.__STATUS_EXTRA: CONFIG.excel_colors.YELLOW,
        }

    # --------------------------------------------------------------
    # Read-only properties
    # --------------------------------------------------------------
    @property
    def toc_path(self) -> Path:
        return self.__toc_path

    @property
    def chunks_path(self) -> Path:
        return self.__chunks_path

    # --------------------------------------------------------------
    # Special methods
    # --------------------------------------------------------------
    def __str__(self) -> str:
        return (
            "ExcelReportGenerator("
            f"toc={self.__toc_path.name}, "
            f"chunks={self.__chunks_path.name}"
            ")"
        )

    def __repr__(self) -> str:
        return (
            "ExcelReportGenerator("
            f"toc_path={self.__toc_path!r}, "
            f"chunks_path={self.__chunks_path!r}, "
            f"output_path={self.output_path!r}"
            ")"
        )

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, ExcelReportGenerator):
            return NotImplemented
        return (
            self.__toc_path == other.__toc_path
            and self.__chunks_path == other.__chunks_path
            and self.output_path == other.output_path
        )

    def __len__(self) -> int:
        return 2

    # --------------------------------------------------------------
    # Context manager
    # --------------------------------------------------------------
    def __enter__(self) -> "ExcelReportGenerator":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> bool:
        return False

    # --------------------------------------------------------------
    # Template method steps
    # --------------------------------------------------------------
    def _validate_inputs(self) -> None:
        if not self.__toc_path.exists():
            raise FileNotFoundError(
                f"ToC file not found: {self.__toc_path}"
            )
        if not self.__chunks_path.exists():
            raise FileNotFoundError(
                f"Chunks file not found: {self.__chunks_path}"
            )

    def _load_data(self) -> Dict[str, Any]:
        return {
            "toc": JSONLHandler.load(self.__toc_path),
            "chunks": JSONLHandler.load(self.__chunks_path),
        }

    def _process_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        toc_map = self._safe_map(data["toc"])
        chunk_map = self._safe_map(data["chunks"])

        all_ids = sorted(
            set(toc_map) | set(chunk_map),
            key=self._sort_key,
        )

        rows: List[Dict[str, Any]] = []

        for sid in all_ids:
            t = toc_map.get(sid)
            c = chunk_map.get(sid)

            rows.append(
                {
                    "section_id": sid,
                    "toc_title": t["title"] if t else None,
                    "toc_page": t["page"] if t else None,
                    "chunk_title": c["title"] if c else None,
                    "chunk_page": c["page"] if c else None,
                    "status": self._resolve_status(t, c),
                }
            )

        return {"rows": rows}

    def _format_output(self, data: Dict[str, Any]) -> pd.DataFrame:
        return pd.DataFrame(data["rows"])

    def _save_report(self, df: pd.DataFrame) -> Path:
        final_output = self._timestamped_filename(
            self.output_path.stem,
            self.output_path.suffix,
        )

        df.to_excel(final_output, index=False, sheet_name="coverage")

        wb = load_workbook(final_output)
        ws = wb["coverage"]

        self._apply_row_colors(ws)
        self._apply_sheet_formatting(ws)
        self._add_summary_sheet(wb, df)

        wb.save(final_output)
        return final_output

    # --------------------------------------------------------------
    # Helpers
    # --------------------------------------------------------------
    @classmethod
    def _resolve_status(cls, toc, chunk) -> str:
        if toc and chunk:
            return cls.__STATUS_MATCHED
        if toc:
            return cls.__STATUS_MISSING
        return cls.__STATUS_EXTRA

    @staticmethod
    def _safe_map(
        data: List[Dict[str, Any]]
    ) -> Dict[str, Dict[str, Any]]:
        return {
            d["section_id"]: d
            for d in data
            if isinstance(d, dict) and "section_id" in d
        }

    @staticmethod
    def _sort_key(section_id: Any) -> List[int]:
        if section_id is None:
            return [9999]

        sid = str(section_id)

        if sid.startswith("FM-"):
            try:
                return [-1, int(sid.split("-")[1])]
            except Exception:
                return [-1, 0]

        try:
            return [int(x) for x in sid.split(".")]
        except Exception:
            return [9999]

    # --------------------------------------------------------------
    # Excel formatting
    # --------------------------------------------------------------
    def _apply_row_colors(self, ws) -> None:
        for row in ws.iter_rows(min_row=2):
            status = row[5].value
            color = self.__colors.get(
                status,
                CONFIG.excel_colors.WHITE,
            )
            fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid",
            )
            for cell in row:
                cell.fill = fill

    @staticmethod
    def _apply_sheet_formatting(ws) -> None:
        ws.auto_filter.ref = "A1:F1"
        ws.freeze_panes = "A2"

        for col in ws.columns:
            values = [
                str(cell.value)
                for cell in col
                if cell.value is not None
            ]
            width = max((len(v) for v in values), default=8)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = width + 2

        for row in ws.iter_rows(min_row=2):
            cell = row[0]
            cell.hyperlink = f"#coverage!A{cell.row}"
            cell.style = "Hyperlink"

    # --------------------------------------------------------------
    # Summary
    # --------------------------------------------------------------
    @staticmethod
    def _add_summary_sheet(wb, df: pd.DataFrame) -> None:
        summary = wb.create_sheet("summary")

        total = len(df)
        matched = int((df["status"] == "MATCHED").sum())
        missing = int((df["status"] == "MISSING_IN_CONTENT").sum())
        extra = int((df["status"] == "EXTRA_IN_CONTENT").sum())
        match_pct = round((matched / total) * 100, 2) if total else 0

        rows = [
            ["Metric", "Value"],
            ["Total Sections", total],
            ["Matched", matched],
            ["Missing", missing],
            ["Extra", extra],
            ["Match %", match_pct],
        ]

        for r in rows:
            summary.append(r)

        for cell in summary[1]:
            cell.font = Font(bold=True)

        chart = BarChart()
        chart.title = "Content Match Summary"

        data_ref = Reference(summary, min_col=2, min_row=3, max_row=5)
        label_ref = Reference(summary, min_col=1, min_row=3, max_row=5)

        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(label_ref)

        summary.add_chart(chart, "D2")
